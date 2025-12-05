from collections import defaultdict
from typing import NamedTuple
from datetime import date

from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.db import SessionLocal
from app.models import Swimmer, PersonalBest, Discipline, Course
from app.api.results import get_best_times_for_age
from .config import DATA_DIR

today = date.today().strftime("%d.%m.%Y")
today_str = date.today().strftime("%Y_%m_%d")
RESULT_FILE = DATA_DIR / f"osobaky{today_str}.xlsx"
RESULT_CLUB_FILE = DATA_DIR / f"club_records_{today_str}.xlsx"


def format_time(ms: int | None) -> str:
    if ms is None:
        return ""
    minutes, rem_ms = divmod(ms, 60_000)
    seconds, milliseconds = divmod(rem_ms, 1_000)
    return f"{minutes:02}:{seconds:02}.{milliseconds // 10:02}"


class PBInfo(NamedTuple):
    name: str
    surname: str
    birth_year: int
    group: str
    discipline: str
    course: str
    time: int
    split_time: bool
    relay_part: bool
    date: date
    location: str


class TimeData:
    def __init__(
        self, time: int, location: str, date: date, split_time: bool, relay_part: bool
    ):
        self.time = format_time(time)
        self.location = location
        self.date = date.strftime("%d.%m.%Y")
        self.split_time = split_time
        self.relay_part = relay_part

    def get_commentary_string(self):
        comment = f"{self.location}, {self.date}"
        if self.split_time:
            comment += " - mezičas"
        if self.relay_part:
            comment += " - štafeta"
        return comment

    def get_time(self):
        return self.time


def discipline_sort_key(disc_key: str) -> tuple:
    """
    Assumes disc_key is like '100 Z 25m' (code + course)
    """

    stroke_order = {
        "Z": 0,
        "P": 1,
        "M": 2,
        "K": 3,
        "O": 4,
    }
    try:
        parts = disc_key.split()  # e.g. ['100', 'Z', '25m']
        distance = int(parts[0])
        stroke_code = parts[1].upper()
        return (stroke_order.get(stroke_code, 99), distance)
    except Exception:
        return (999, 999)


def calc_table_range(data):
    """
    Calculate the range of a table based on the data provided.
    :param data: List of lists containing the data for the table.
    :return: String representing the range of the table (e.g., "A1:D10").
    """
    num_rows = len(data)
    num_cols = len(data[0])
    start_cell = "A1"
    end_cell = f"{chr(65 + num_cols - 1)}{num_rows}"
    return f"{start_cell}:{end_cell}"


def load_personal_bests():
    """Fetches personal bests from the database.

    Returns:
        list[PBInfo]: List of personal bests with swimmer details.
    """

    db: Session = SessionLocal()

    pbs: list[PBInfo] = [
        PBInfo(*row)
        for row in db.query(
            Swimmer.name,
            Swimmer.surname,
            Swimmer.birth_year,
            Swimmer.group,
            Discipline.code,
            Course.type,
            PersonalBest.time,
            PersonalBest.split_time,
            PersonalBest.relay_part,
            PersonalBest.date,
            PersonalBest.competition_location,
        )
        .join(PersonalBest.swimmer)
        .join(PersonalBest.swimmer)
        .join(PersonalBest.discipline)
        .join(PersonalBest.course)
        .order_by(Swimmer.surname, Swimmer.name, Swimmer.id)
        .all()
    ]

    db.close()

    return pbs


def get_all_swimmers():
    db: Session = SessionLocal()

    groups = ["Z1", "Z2", "P1", "veteran"]

    swimmers = (
        db.query(Swimmer)
        .filter(Swimmer.group.in_(groups))
        .order_by(Swimmer.surname, Swimmer.name, Swimmer.id)
        .all()
    )

    db.close()

    return swimmers


def get_all_disciplines():
    db: Session = SessionLocal()

    disciplines = [row[0] for row in db.query(Discipline.code).all()]

    db.close()

    return disciplines


def init_data():
    """Initialize the data structure for storing personal bests."""
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    all_swimmers = get_all_swimmers()
    all_disciplines = get_all_disciplines()

    for swimmer in all_swimmers:
        swimmer_key = (swimmer.name, swimmer.surname, swimmer.birth_year)
        for course in ["SCM", "LCM"]:
            for discipline in all_disciplines:
                data[course][swimmer.group][swimmer_key][discipline] = None

    return data


header_fill = PatternFill(fill_type="solid", start_color="B00D08")
white_font = Font(bold=True, color="FFFFFF")

red_font = Font(color="FF0000")

# Color setup
gray_fill_1 = PatternFill("solid", start_color="8D8D8D")
gray_fill_2 = PatternFill("solid", start_color="D6D6D6")
split_fill = PatternFill("solid", start_color="FFCC00")
relay_fill = PatternFill("solid", start_color="FF9900")
red_fill_1 = PatternFill("solid", start_color="732B29")
red_fill_2 = PatternFill("solid", start_color="953735")
title_fill = PatternFill("solid", start_color="000000")
name_fill = PatternFill("solid", fgColor="DE5A72")
white_font = Font(color="FFFFFF")
bold_font = Font(bold=True)
title_font = Font(bold=True, size=20, color="FFFFFF")
group_header_font = Font(bold=True, size=16)
group_header_fill = PatternFill("solid", start_color="B00D08")
border = Border(*[Side(style="thin", color="000000")] * 4)

# Stroke sort order by code
stroke_order = {"M": 0, "Z": 1, "P": 2, "K": 3, "O": 4}
group_order = {"Z1": 0, "Z2": 1, "P1": 2, "veteran": 3, None: 99}
group_display_name = {
    "Z1": "Z1",
    "Z2": "Z2",
    "P1": "P1",
    "veteran": "bývalí",
    None: "Nezařazení",
}


def generate_records(wb: Workbook):
    """Generate a record sheet with the best times for each discipline."""
    ws = wb.create_sheet(title="Records")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 8

    # Header row
    headers = ["Disciplína", "Z1", "Z2", "P1", "Bývalí"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Fetch records from the database
    db: Session = SessionLocal()
    records = (
        db.query(
            Discipline.code,
            Swimmer.group,
            PersonalBest.time,
            PersonalBest.date,
            PersonalBest.competition_location,
        )
        .join(PersonalBest.swimmer)
        .join(PersonalBest.discipline)
        .order_by(Discipline.code, Swimmer.group, PersonalBest.time)
        .all()
    )
    db.close()

    # Organize records by discipline and group
    record_data = defaultdict(lambda: defaultdict(list))
    for code, group, time, record_date, location in records:
        record_data[code][group].append((time, record_date, location))

    current_row = 2
    for code, groups in sorted(
        record_data.items(), key=lambda x: stroke_order.get(x[0][0], 99)
    ):
        ws.cell(row=current_row, column=1, value=code).font = bold_font

        for col_index, group in enumerate(["Z1", "Z2", "P1", "veteran"], start=2):
            if group in groups:
                # Get the best time for this group
                time_data = groups[group][0]
                time_str = format_time(time_data[0])
                comment_str = f"{time_data[1]} - {time_data[2]}"
                cell = ws.cell(row=current_row, column=col_index, value=time_str)
                cell.comment = Comment(comment_str)
            else:
                cell = ws.cell(row=current_row, column=col_index)

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = gray_fill_1 if current_row % 2 == 0 else gray_fill_2


def generate_excel_from_pbs(pbs: list[PBInfo], filename: str = RESULT_FILE):
    # Organize data by course type -> group -> swimmer_id -> {discipline: PBInfo}
    all_discipline_keys = defaultdict(set)

    structured_data = init_data()

    group_order = {
        "Z1": 0,
        "Z2": 1,
        "P1": 2,
        "veteran": 3,
        None: 99,  # in case some swimmers have no group
    }
    group_display_name = {
        "Z1": "Z1",
        "Z2": "Z2",
        "P1": "P1",
        "veteran": "bývalí",
        None: "Nezařazení",
    }

    for pb in pbs:
        course = pb.course
        swimmer_key = (pb.name, pb.surname, pb.birth_year)
        group = pb.group
        disc_key = pb.discipline

        structured_data[course][group][swimmer_key][disc_key] = pb
        all_discipline_keys[course].add(disc_key)

    discipline_keys = sorted(all_discipline_keys["SCM"], key=discipline_sort_key)
    headers = ["Jméno", "Ročník"] + [disc for disc in discipline_keys]
    total_columns = len(headers)

    wb = Workbook()

    for i, (course, groups) in enumerate(structured_data.items()):

        def _course_map(course):
            return "25m" if course == "SCM" else "50m"

        ws = wb.active if i == 0 else wb.create_sheet(title=course)
        ws.title = _course_map(course)
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 8

        current_row = 1
        start_col = get_column_letter(1)
        end_col = get_column_letter(total_columns)
        ws.merge_cells(f"{start_col}{current_row}:{end_col}{current_row}")
        cell = ws.cell(
            row=current_row,
            column=1,
            value=f"Osobní rekordy {'25' if course == 'SCM' else '50'} m - aktualizováno {today}",
        )
        cell.font = title_font
        cell.fill = title_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        for group_name in sorted(groups.keys(), key=lambda g: group_order.get(g, 99)):
            group_label = group_display_name.get(group_name, group_name)

            # Group header row (merged)
            start_col = get_column_letter(1)
            end_col = get_column_letter(total_columns)
            ws.merge_cells(f"{start_col}{current_row}:{end_col}{current_row}")
            cell = ws.cell(row=current_row, column=1, value=f"{group_label}")
            cell.font = group_header_font
            cell.fill = group_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

            # Header row
            for col_index, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_index, value=header)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = header_fill
                cell.font = bold_font
                cell.border = border

            current_row += 1

            # Swimmer rows
            swimmers = groups[group_name]
            for swimmer_key, pb_map in swimmers.items():
                name, surname, year = swimmer_key
                row_data = [f"{surname} {name}", year]
                for disc in discipline_keys:
                    pb = pb_map.get(disc)
                    if pb:
                        time_data = TimeData(
                            pb.time, pb.location, pb.date, pb.split_time, pb.relay_part
                        )
                        row_data.append(time_data)
                    else:
                        row_data.append(None)

                fill = gray_fill_1 if current_row % 2 == 0 else gray_fill_2

                for col_index, cell_data in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_index)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True)
                    cell.fill = fill

                    if col_index in (1, 2):  # name, surname
                        cell.fill = red_fill_1 if current_row % 2 == 0 else red_fill_2
                        cell.font = bold_font
                        alignement = "left" if col_index == 1 else "center"
                        cell.alignment = Alignment(
                            horizontal=alignement, vertical="center"
                        )
                        cell.value = cell_data

                    if isinstance(cell_data, TimeData):
                        cell.value = cell_data.get_time()
                        cell.comment = Comment(cell_data.get_commentary_string(), ".")
                        if cell_data.split_time:
                            cell.fill = split_fill
                        if cell_data.relay_part:
                            cell.fill = relay_fill
                ws.row_dimensions[current_row].height = 30
                current_row += 1

            ws.row_dimensions[current_row].height = 10
            current_row += 1  # spacer between groups

    # generate_records(wb)

    wb.save(filename)


def print_records(records, gender, course_length):
    """Prints the records in a formatted way."""
    print("\n\n\n==========================================")
    print(f"Rekordy {course_length}m - {'muži' if gender == 'M' else 'ženy'}")
    for record in records:
        discipline, time, name, surname, course_type, course_length, date = record
        print(f"{discipline:25} {format_time(time):10} {name:10} {surname:10}  {date}")


def debug():
    pbs = load_personal_bests()

    generate_excel_from_pbs(pbs)

    # db: Session = SessionLocal()
    # men_short = best_records_query(db, gender="M", course_length=25)

    # women_short = best_records_query(db, gender="Z", course_length=25)

    # men_long = best_records_query(db, gender="M", course_length=50)

    # women_long = best_records_query(db, gender="Z", course_length=50)

    # limit = 20
    # top = (db.query(
    #     Swimmer.name,
    #     Swimmer.surname,
    #     Discipline.title,
    #     PersonalBest.points
    # ).join(Swimmer, PersonalBest.swimmer_id == Swimmer.id)
    #     .join(Discipline, PersonalBest.discipline_id == Discipline.id)
    #     .order_by(PersonalBest.points.desc())
    #     .limit(limit)
    #     .all()
    # )
    #
    # print(top)
    #
    # db.close()


def debugo():
    # Base.metadata.create_all(bind=engine)

    db: Session = SessionLocal()

    print("are we here?")

    ## get the veteran swimmers, sorted by surname
    veterans = (
        db.query(Swimmer)
        .filter(Swimmer.group == "veteran")
        .order_by(Swimmer.surname, Swimmer.name)
        .all()
    )

    # print output in yml format,
    # like
    # - surname: SURNAME
    #   name: NAME
    for v in veterans:
        print(f'- surname: "{v.surname}"\n  name: "{v.name}"')

    # sync = db.query(ApiSync).first()
    #
    # if not sync:
    #     sync = ApiSync(id=1)
    #     db.add(sync)
    # db.commit()
    db.close()


def create_club_records_excel():
    """Create club records Excel file with categories as columns"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Club Records"

    # Define disciplines structure
    disciplines = [
        ("M", [50, 100, 200]),  # Butterfly
        ("Z", [50, 100, 200]),  # Backstroke
        ("P", [50, 100, 200]),  # Breaststroke
        ("K", [50, 100, 200, 400, 800, 1500]),  # Freestyle
        ("O", [100, 200, 400]),  # Individual Medley
    ]

    stroke_names = {"M": "M", "Z": "Z", "P": "P", "K": "VZ", "O": "O"}

    # Age categories
    ages = list(range(9, 15)) + [18, 99]  # 9-14, Dorost (15-18), Absolutní (25)

    # Styles
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(
        start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"
    )
    subheader_font = Font(bold=True, size=10)
    discipline_fill = PatternFill(
        start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
    )
    discipline_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    # Stronger borders for separating categories and disciplines
    thick_right_border = Border(
        left=Side(style="thin"),
        right=Side(style="medium"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    thick_bottom_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="medium"),
    )
    thick_bottom_right_border = Border(
        left=Side(style="thin"),
        right=Side(style="medium"),
        top=Side(style="thin"),
        bottom=Side(style="medium"),
    )

    # Create headers
    # Row 1: Title row spanning entire table
    total_cols = 1 + (len(ages) * 3)  # 1 discipline column + 3 columns per age
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(
        row=1, column=1, value=f"Klubové rekordy PKBoh - aktualizováno {today}"
    )
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(
        start_color="B00D08", end_color="B00D08", fill_type="solid"
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = border
    ws.row_dimensions[1].height = 30

    # Row 2: Age categories
    col = 2
    for age in ages:
        age_title = f"{age}letí" if age < 15 else "Dorost" if age < 19 else "Absolutní"
        cell = ws.cell(row=2, column=col, value=age_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        col += 3

    # Row 3: Sub-headers
    ws.cell(row=3, column=1, value="Disciplína")
    ws.cell(row=3, column=1).font = discipline_font
    ws.cell(row=3, column=1).fill = discipline_fill
    ws.cell(row=3, column=1).border = border
    ws.cell(row=3, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    col = 2
    for _ in ages:
        for header in ["Jméno", "Čas", "Termín"]:
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            col += 1

    # Set column widths
    ws.column_dimensions["A"].width = 18  # Narrower discipline column
    for col_idx in range(2, col):
        letter = get_column_letter(col_idx)
        if (col_idx - 2) % 3 == 0:
            ws.column_dimensions[letter].width = 20
        elif (col_idx - 2) % 3 == 1:
            ws.column_dimensions[letter].width = 12
        else:
            ws.column_dimensions[letter].width = 15

    # Populate data
    current_row = 4  # Start from row 4 now (after title, age headers, and sub-headers)
    db = SessionLocal()

    try:
        for stroke_idx, (stroke_code, distances) in enumerate(disciplines):
            for distance in distances:
                for sex in ["female", "male"]:
                    discipline_code = f"{distance} {stroke_code}"
                    gender_label = "Muži" if sex == "male" else "Ženy"
                    stroke_name = stroke_names.get(stroke_code, stroke_code)
                    discipline_label = f"{distance} {stroke_name} - {gender_label}"

                    # Determine if this is the last row of a discipline group (last gender of last distance)
                    is_last_distance = distance == distances[-1]
                    is_last_gender = sex == "male"
                    is_discipline_boundary = is_last_distance and is_last_gender

                    cell = ws.cell(row=current_row, column=1, value=discipline_label)
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    # Apply thick bottom border for discipline boundaries
                    if is_discipline_boundary and stroke_idx < len(disciplines) - 1:
                        cell.border = thick_bottom_border
                    else:
                        cell.border = border

                    col = 2
                    for age_idx, age in enumerate(ages):
                        is_last_age_col = age_idx == len(ages) - 1

                        try:
                            records = get_best_times_for_age(
                                db=db,
                                discipline_code=discipline_code,
                                course_length=25,
                                sex=sex,
                                max_age=age,
                                limit=1,
                                unique_swimmers=True,
                            )

                            if records and len(records) > 0:
                                record = records[0]
                                # Name
                                name_cell = ws.cell(
                                    row=current_row,
                                    column=col,
                                    value=f"{record.surname} {record.name}",
                                )
                                name_cell.alignment = Alignment(
                                    horizontal="left", vertical="center"
                                )

                                # Time
                                time_cell = ws.cell(
                                    row=current_row,
                                    column=col + 1,
                                    value=format_time(record.time),
                                )
                                time_cell.alignment = Alignment(
                                    horizontal="center", vertical="center"
                                )

                                # Place and Date (two lines in same cell)
                                date_formatted = (
                                    record.date.strftime("%d.%m.%Y")
                                    if isinstance(record.date, date)
                                    else str(record.date)
                                )
                                place_cell = ws.cell(
                                    row=current_row,
                                    column=col + 2,
                                    value=f"{record.competition_location}\n{date_formatted}",
                                )
                                place_cell.alignment = Alignment(
                                    horizontal="center",
                                    vertical="center",
                                    wrap_text=True,
                                )

                                # Apply borders - thick right border at age category boundaries
                                if (
                                    is_discipline_boundary
                                    and stroke_idx < len(disciplines) - 1
                                ):
                                    # Thick bottom border
                                    if is_last_age_col:
                                        name_cell.border = thick_bottom_right_border
                                        time_cell.border = thick_bottom_right_border
                                        place_cell.border = thick_bottom_right_border
                                    else:
                                        name_cell.border = thick_bottom_border
                                        time_cell.border = thick_bottom_border
                                        place_cell.border = thick_bottom_right_border
                                else:
                                    # Normal borders with thick right at age boundaries
                                    if is_last_age_col:
                                        name_cell.border = thick_right_border
                                        time_cell.border = thick_right_border
                                        place_cell.border = thick_right_border
                                    else:
                                        name_cell.border = border
                                        time_cell.border = border
                                        place_cell.border = thick_right_border
                            else:
                                # Empty cells
                                for offset in range(3):
                                    empty_cell = ws.cell(
                                        row=current_row, column=col + offset
                                    )
                                    # Apply appropriate borders
                                    if (
                                        is_discipline_boundary
                                        and stroke_idx < len(disciplines) - 1
                                    ):
                                        # At discipline boundary
                                        if offset == 2:
                                            # Last column of age category (Termín)
                                            empty_cell.border = (
                                                thick_bottom_right_border
                                            )
                                        else:
                                            # Other columns (Name, Time)
                                            empty_cell.border = thick_bottom_border
                                    else:
                                        # Not at discipline boundary
                                        if offset == 2:
                                            # Last column of age category (Termín)
                                            empty_cell.border = thick_right_border
                                        else:
                                            # Other columns
                                            empty_cell.border = border

                        except Exception as e:
                            print(f"Error: {discipline_label}, age {age}: {e}")
                            for offset in range(3):
                                empty_cell = ws.cell(
                                    row=current_row, column=col + offset
                                )
                                # Apply appropriate borders even on error
                                if (
                                    is_discipline_boundary
                                    and stroke_idx < len(disciplines) - 1
                                ):
                                    # At discipline boundary
                                    if offset == 2:
                                        # Last column of age category (Termín)
                                        empty_cell.border = thick_bottom_right_border
                                    else:
                                        # Other columns (Name, Time)
                                        empty_cell.border = thick_bottom_border
                                else:
                                    # Not at discipline boundary
                                    if offset == 2:
                                        # Last column of age category (Termín)
                                        empty_cell.border = thick_right_border
                                    else:
                                        # Other columns
                                        empty_cell.border = border

                        col += 3

                    current_row += 1

    finally:
        db.close()

    ws.freeze_panes = "B4"  # Freeze first 3 rows (title, age categories, sub-headers) and first column

    wb.save(RESULT_CLUB_FILE)
    print(f"Saved to {RESULT_CLUB_FILE}")


if __name__ == "__main__":
    # debugi()
    create_club_records_excel()
    # debug()
    # debugo()
